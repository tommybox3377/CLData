from bs4 import BeautifulSoup as bs
import openpyxl

from urllib.parse import urlparse
import requests
import datetime
import re
import math
import os

###########################################################################################################
### Set this stuff
# Make an ***EMPTY** folder named "CraigsListSearch" and place its path as shown below (replace "C:\Users\twmar\OneDrive\Desktop" with your location)
folder_location = r"C:\Users\twmar\OneDrive\Desktop"

# over write the following information for all the cars you want to search:
raw_car_data = [
    # make, model, (trims/submodels)
    ("mitsubishi", "lancer", ("es", "gsr", "mr", "gts", "gt"))
]

# set owners to true if you want to search for cars being sold by the car owner
# set dealers to true if you want to search for cars being sold by dealers
# can set either or both to True
owners = True
dealers = True

# set your latitude and longitude, you can get this from GoogleMaps by right clicking on a location and clicking "What's Here?" the latitude and longitude will show up on the bottom. Central Park in New York, New York USA is currently entered for example
my_latitude = 40.782316
my_longitude = -73.965703


# Let the rest of the code do magic stuff
###########################################################################################################
class Car:
    def __init__(self, make, model, trims):
        self.make = make
        self.model = model
        self.trims = trims

c = {}
cars = []

for i in range(len(raw_car_data)):
    c[i] = Car(raw_car_data[i][0], raw_car_data[i][1], (raw_car_data[i][2]))

for x in c.values():
    cars.append(x)

### make all files
for path in os.listdir(folder_location + "/CraigsListSearch"):
    full_path = os.path.join(folder_location + "/CraigsListSearch", path)
    if os.path.isfile(full_path) and "CraigsListSearch/Data.xlsx" in full_path or "CraigsListSearch/Cites.txt" in full_path or "CraigsListSearch/Log.txt" in full_path:
        os.remove(full_path)
data_file = folder_location + "/CraigsListSearch/Data.xlsx"
CL_cites_file = folder_location + "/CraigsListSearch/Cites.txt"
log_file = folder_location + "/CraigsListSearch/Log.txt"
URLs_to_scrape = []
with open(CL_cites_file, "w") as f:
    pass
with open(log_file, "w") as f:
    pass


### Logger
# Error = when the approximate location of the crash is known
def LogError(error):
    f = open(log_file, "a")
    f.write("\n" + datetime.datetime.now().strftime("%m/%d/%Y, %H:%M:%S") + " ERROR: " + error)
    f.close()


# Note = when the data to be inserted is as good as its going to get
def LogNote(note):
    f = open(log_file, "a")
    f.write("\n" + datetime.datetime.now().strftime("%m/%d/%Y, %H:%M:%S") + " Note: " + note)
    f.close()


# Crash = Catchall for the entire scrape
def LogCrash(error):
    f = open(log_file, "a")
    f.write("\n" + datetime.datetime.now().strftime("%m/%d/%Y, %H:%M:%S") + " CRASH: " + error)
    f.close()

car_attr = [
    "URL", "DatetimePosted", "DaysSincePosted", "Latitude", "Longitude", "NumOfPics", "CarCondition",
    "Title", "Year", "Trim", "Fuel", "Odometer", "TitleStatus", "Transmission", "Cylinders", "Drive",
    "PaintColor", "Size", "Type", "Price", "PostLength", "CLCity", "NumOfAttributes", "PostID",
    "PriceHistory", "Make", "Model", "Website", "CarAge", "DateUpdated", "DaysSinceUpdated", "Distance"
]

workbook = openpyxl.Workbook()
sheet = workbook.active
for i, attr in enumerate(car_attr):
    sheet.cell(row=1, column=i+1, value=attr)
workbook.save(filename=data_file)


### CL
def calc_dist(point1, point2):
    R = 6373.0
    lat1 = math.radians(point1[0])
    lon1 = math.radians(point1[1])
    lat2 = math.radians(point2[0])
    lon2 = math.radians(point2[1])
    dlon = lon2 - lon1
    dlat = lat2 - lat1
    a = math.sin(dlat / 2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return int((R * c)*.621371)


def get_all_cl_cities_urls():
    urls = []
    response = requests.get(r"https://www.craigslist.org/about/sites")
    content = response.content
    soup = bs(content, "html.parser")
    all_in_US = soup.find_all("div", {"class": "colmask"})[0]  # 0 index is just for how craigslist breaks up the urls by location
    for link in all_in_US.find_all('a'):
        urls.append(link.get('href'))
    return urls


with open(CL_cites_file, "w") as f:
    for u in get_all_cl_cities_urls():
        f.write(u + "\n")


def get_all_urls(url, make, model):
    if owners and dealers:
        sold_by = "cta"
    elif owners:
        sold_by = "cto"
    elif dealers:
        sold_by = "ctd"
    else:
        return "Enter True for either or both owners/dealers vars"

    url_string = "%ssearch/%s?auto_make_model=%s+%s" % (url, sold_by, make, model)
    response = requests.get(url_string)
    content = response.content
    soup = bs(content, "html.parser")
    all = soup.find_all("li", {"class": "result-row"})
    for results in all:
        for link in results.find_all('a', {"class": "result-title hdrlnk"}):
            car_url = (link.get('href'))
            if url in car_url:
                URLs_to_scrape.append(car_url)


print("getting all car URLs")

with open(CL_cites_file) as f:
    for c in cars:
        for city in f.readlines():
            get_all_urls(city.strip(), c.make.strip(), c.model.strip())

print("searching all URLS")


def check_if_listing_exists(url):
    try:
        response = requests.get(url)
        if response.status_code != 200:
            if response.status_code == 404:
                LogNote(f"{url} had a 404 response code")
            else:
                LogError(f"{url} did not have a 200 response code")
            return False
        if "This posting has been flagged for removal." in response.text:
            LogNote(f"{url} was flagged for removal")
            return False
        if "This posting has expired" in response.text:
            LogNote(f"{url} has expired")
            return False
        if "This posting has been deleted by its author." in response.text:
            LogNote(f"{url} has been deleted by its author")
            return False
        return True
    except:
        LogError(f"{url} failed try block in check_listing_exists")
        return False


def first_search(url, car):
    if not check_if_listing_exists(url):
        return {}
    try:
        now = datetime.datetime.now()
        results = {
            "URL": url,
            "DatetimePosted": None,
            "DaysSincePosted": None,
            "Latitude": None,
            "Longitude": None,
            "NumOfPics": None,
            "CarCondition": None,
            "Title": None,
            "Year": None,
            "Trim": None,
            "Fuel": None,
            "Odometer": None,
            "TitleStatus": None,
            "Transmission": None,
            "Cylinders": None,
            "Drive": None,
            "PaintColor": None,
            "Size": None,
            "Type": None,
            "Price": None,
            "PostLength": None,
            "CLCity": urlparse(url)[1].split(".")[0],
            "NumOfAttributes": None,
            "PostID": None,
            "PriceHistory": None,
            "Make": car.make,
            "Model": car.model,
            "Website": "craigslist.org",
            "CarAge": None,
            "DateUpdated": None,
            "DaysSinceUpdated": None,
            "Distance": 0
        }
        response = requests.get(url)
        content = response.text
        soup = bs(content, "html.parser")

        info = soup.find_all("p", {"class": "postinginfo reveal"})
        for date in info:
            if "posted:" in date.get_text():
                results["DatetimePosted"] = datetime.datetime.strptime(date.find("time").get_text(), "%Y-%m-%d %H:%M")
                results["DaysSincePosted"] = (now - results["DatetimePosted"]).days
            elif "updated:" in date.get_text():
                results["DateUpdated"] = datetime.datetime.strptime(date.find("time").get_text(), "%Y-%m-%d %H:%M")
                results["DaysSinceUpdated"] = (now - results["DateUpdated"]).days

        info = soup.find("div", {"id": "map"})
        if info:
            results["Latitude"] = info.get("data-latitude")
            results["Longitude"] = info.get("data-longitude")
            results["Distance"] = calc_dist((float(results["Latitude"]), float(results["Longitude"])), (my_latitude, my_longitude))

        info = soup.find("div", {"id": "thumbs"})
        if info:
            results["NumOfPics"] = len(info.find_all("a", {"class": "thumb"}))
        else:
            results["NumOfPics"] = 0

        try:
            info = soup.find_all("p", {"class": "attrgroup"})

            results["Title"] = (info[0].get_text().strip())
            atts = info[1].find_all("span")
            results["NumOfAttributes"] = len(atts)
            for att in atts:
                atr = att.contents[0]
                try:
                    val = att.find("b").get_text()
                except:
                    if "delivery available" not in atr.lower() and "cryptocurrency" not in atr.lower():
                        LogNote(f"attribute only had one index at {url}")
                if "condition" in atr:
                    results["CarCondition"] = val
                elif "fuel" in atr:
                    results["Fuel"] = val
                elif "odometer" in atr:
                    results["Odometer"] = val
                elif "title status" in atr:
                    results["TitleStatus"] = val
                elif "transmission" in atr:
                    results["Transmission"] = val
                elif "cylinders" in atr:
                    try:
                        results["Cylinders"] = re.search("\d+", val)[0]
                    except:
                        if "other" not in val:
                            LogError(f"could not read cylinders at {url}")
                elif "drive" in atr:
                    results["Drive"] = val
                elif "paint color" in atr:
                    results["PaintColor"] = val
                elif "size" in atr:
                    results["Size"] = val
                elif "type" in atr:
                    results["Type"] = val
                elif "vin" in atr.lower():
                    pass
                elif "delivery available" in atr.lower():
                    pass
                elif "cryptocurrency" in atr.lower():
                    pass
                else:
                    LogNote(f"Attribute {atr} was not found in dictionary for {url}")
        except Exception as e:
            LogNote("Attribute search error for " + url + str(e))

        info = soup.find("section", {"id": "postingbody"})
        if info:
            text = info.get_text()
            l = len(info.get_text())
            if "QR Code Link to This Post" in text:
                l -= 25
            results["PostLength"] = l

        info = soup.find("span", {"class": "price"})
        if info:
            price = re.search("\$*(\d+)", info.get_text())
            results["Price"] = (price[1])

        info = soup.find_all("p", {"class": "postinginfo"})
        for i in info:
            if "post id" in i.get_text():
                results["PostID"] = re.search("\d+", i.get_text())[0]

        if results["Title"]:
            title = results["Title"]
            results["Year"] = re.search("\d+", title)[0]
            #TODO check that year is in 4 digit format
            results["CarAge"] = now.year - int(results["Year"])

        if results["Price"]:
            results["PriceHistory"] = f"{now.strftime('%m/%d/%Y, %H:%M:%S')}${str(results['Price'])}"

        info = soup.find("span", {"id": "titletextonly"})
        for trim in car.trims:
            if re.search(f"[^a-z]{trim}[^a-z]*", info.get_text().lower()) or re.search(f"[^a-z]{trim}[^a-z]*", results["Title"]):
                results["Trim"] = trim
                break
    except Exception as e:
        LogCrash(f"{url} crashed during scraping of html response {str(e)}")

    # for x, y in results.items():
    #     print(x, y)

    return results


workbook = openpyxl.load_workbook(data_file)
sheet = workbook.active
for c in cars:
    for u in URLs_to_scrape:
        sheet.append(list(first_search(u, c).values()))
workbook.save(filename=data_file)

print("done")
